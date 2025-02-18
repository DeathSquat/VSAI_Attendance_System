import cv2
import face_recognition
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Paths
students_path = r"3. Attendance System/Students"
attendance_file = r"3. Attendance System/Attendance.xlsx"

# Load student images and encode their faces
student_encodings = []
student_names = []

for filename in os.listdir(students_path):
    if filename.endswith(".jpg") or filename.endswith(".png"):
        student_image = face_recognition.load_image_file(os.path.join(students_path, filename))
        encoding = face_recognition.face_encodings(student_image)

        # Check if a face was detected and successfully encoded
        if len(encoding) > 0:
            student_encodings.append(encoding[0])
            student_names.append(os.path.splitext(filename)[0])

# Function to initialize the monthly sheet
def initialize_month_sheet():
    now = datetime.now()
    month_year = now.strftime('%B-%Y')  # e.g., "January-2025"

    if not os.path.exists(attendance_file):
        workbook = Workbook()
        workbook.save(attendance_file)

    workbook = load_workbook(attendance_file)

    # Check if the sheet for the current month already exists
    if month_year not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=month_year)
        sheet.append(["Student Name"] + [f"Day {i}" for i in range(1, 32)])  # Max days in a month

        # Initialize all students with "Absent" (red cells)
        for student_name in student_names:
            row = [student_name] + ["Absent"] * 31
            sheet.append(row)

        # Style the sheet (make "Absent" cells red)
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=32, max_row=len(student_names) + 1):
            for cell in row:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        workbook.save(attendance_file)
        print(f"New sheet '{month_year}' initialized.")
    else:
        print(f"Sheet '{month_year}' already exists.")

# Function to initialize the summary sheet
def initialize_summary_sheet():
    workbook = load_workbook(attendance_file)

    # Check if the "Attendance Summary" sheet exists
    if "Attendance Summary" not in workbook.sheetnames:
        summary_sheet = workbook.create_sheet(title="Attendance Summary")
        summary_sheet.append(["Student Name", "Total Attendance"] + [f"{month}" for month in workbook.sheetnames if month != "Attendance Summary"])
        
        # Initialize rows for each student
        for student_name in student_names:
            summary_sheet.append([student_name, 0] + [0] * (len(workbook.sheetnames) - 1))

        workbook.save(attendance_file)
        print("Attendance Summary sheet initialized.")
    else:
        print("Attendance Summary sheet already exists.")

# Function to update summary sheet
def update_summary_sheet(student_name, month_year):
    workbook = load_workbook(attendance_file)
    summary_sheet = workbook["Attendance Summary"]

    # Find the row of the student
    student_row = None
    for row in range(2, summary_sheet.max_row + 1):
        if summary_sheet.cell(row, 1).value == student_name:
            student_row = row
            break

    # If student is not found, add them
    if student_row is None:
        student_row = summary_sheet.max_row + 1
        summary_sheet.cell(student_row, 1).value = student_name
        summary_sheet.cell(student_row, 2).value = 0

    # Update total attendance for the month
    month_column = None
    for col in range(3, summary_sheet.max_column + 1):
        if summary_sheet.cell(1, col).value == month_year:
            month_column = col
            break

    if month_column is None:  # Add a new column for the current month
        month_column = summary_sheet.max_column + 1
        summary_sheet.cell(1, month_column).value = month_year

    # Increment attendance count for the current month and total
    current_month_attendance = summary_sheet.cell(student_row, month_column).value or 0
    summary_sheet.cell(student_row, month_column).value = current_month_attendance + 1

    total_attendance = summary_sheet.cell(student_row, 2).value or 0
    summary_sheet.cell(student_row, 2).value = total_attendance + 1

    workbook.save(attendance_file)
    print(f"Attendance of {student_name} has been Marked.")

# Function to mark attendance for a student
def mark_attendance(student_name):
    now = datetime.now()
    day = now.day
    month_year = now.strftime('%B-%Y')  # e.g., "January-2025"

    workbook = load_workbook(attendance_file)

    # Ensure the sheet for the current month exists
    if month_year not in workbook.sheetnames:
        initialize_month_sheet()

    sheet = workbook[month_year]

    # Find the student's row
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == student_name:
            cell = sheet.cell(row, day + 1)
            if cell.value != "Present":  # Only mark if not already "Present"
                cell.value = "Present"
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for Present
                workbook.save(attendance_file)
                update_summary_sheet(student_name, month_year)  # Update summary
            return

    # If the student is not in the sheet, add them
    new_row = sheet.max_row + 1
    sheet.cell(new_row, 1).value = student_name
    for col in range(2, 33):  # Initialize all days as "Absent"
        cell = sheet.cell(new_row, col)
        cell.value = "Absent"
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for Absent
    sheet.cell(new_row, day + 1).value = "Present"  # Mark the current day as Present
    sheet.cell(new_row, day + 1).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    workbook.save(attendance_file)
    update_summary_sheet(student_name, month_year)  # Update summary

# Function to display options menu
def show_options():
    while True:
        print("\n--- Options Menu ---")
        print("1. View total attendance for a student")
        print("2. Remove a student's attendance record")
        print("3. Exit options menu")
        
        choice = input("Enter your choice (1-3): ")
        
        if choice == "1":
            view_total_attendance()
        elif choice == "2":
            remove_student_record()
        elif choice == "3":
            break
        else:
            print("Invalid choice. Please try again.")

# Function to view total attendance of a student
def view_total_attendance():
    student_name = input("Enter the student's name: ")
    workbook = load_workbook(attendance_file)
    if "Attendance Summary" not in workbook.sheetnames:
        print("Attendance Summary sheet does not exist.")
        return

    summary_sheet = workbook["Attendance Summary"]
    for row in range(2, summary_sheet.max_row + 1):
        if summary_sheet.cell(row, 1).value == student_name:
            total_attendance = summary_sheet.cell(row, 2).value or 0
            print(f"Total Attendance for {student_name}: {total_attendance}")
            return

    print(f"Student {student_name} not found in the records.")

# Function to remove a student's attendance record
def remove_student_record():
    student_name = input("Enter the student's name to remove: ")
    workbook = load_workbook(attendance_file)
    
    # Remove from monthly sheets
    for sheet_name in workbook.sheetnames:
        if sheet_name != "Attendance Summary":
            sheet = workbook[sheet_name]
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, 1).value == student_name:
                    sheet.delete_rows(row, 1)
                    print(f"Removed {student_name} from sheet {sheet_name}.")
                    break

    # Remove from summary sheet
    if "Attendance Summary" in workbook.sheetnames:
        summary_sheet = workbook["Attendance Summary"]
        for row in range(2, summary_sheet.max_row + 1):
            if summary_sheet.cell(row, 1).value == student_name:
                summary_sheet.delete_rows(row, 1)
                print(f"Removed {student_name} from Attendance Summary.")
                break

    workbook.save(attendance_file)
    print(f"Attendance record for {student_name} has been removed.")

# Initialize the monthly sheet and summary sheet
initialize_month_sheet()
initialize_summary_sheet()

# Initialize the webcam
cap = cv2.VideoCapture(0)

# Track students who have already been marked
marked_students = set()

while True:
    # Capture a frame from the webcam
    ret, frame = cap.read()

    if not ret:
        break

    # Resize frame for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)

    # Detect face locations and encodings in the current frame
    face_locations = face_recognition.face_locations(small_frame)
    face_encodings = face_recognition.face_encodings(small_frame, face_locations)

    # Loop through each detected face
    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        # Compare face with the loaded student encodings
        matches = face_recognition.compare_faces(student_encodings, face_encoding)
        name = "Unknown"

        # If a match is found, use the student's name
        if True in matches:
            match_index = matches.index(True)
            name = student_names[match_index]

            # Mark attendance if not already marked
            if name not in marked_students:
                mark_attendance(name)
                marked_students.add(name)

        # Draw a rectangle around the detected face and label it
        top *= 2
        right *= 2
        bottom *= 2
        left *= 2

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    # Display the frame
    cv2.imshow("Student Attendance System", frame)

    # Check for key presses
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):  # Quit the program
        break
    elif key == ord('o'):  # Show options menu
        show_options()

# Release the webcam and close the OpenCV window
cap.release()
cv2.destroyAllWindows()